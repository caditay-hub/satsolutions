#!/usr/bin/env python3
"""
prepare.py — извлекает из прайс-листа Dahua правильные фото и карту товаров.

Читает .xlsx, в котором фото встроены в колонку B и растянуты на несколько строк.
Привязывает каждое фото к товару по ВЕРТИКАЛЬНОМУ ЦЕНТРУ изображения (с учётом высот
строк и смещений EMU) — это исправляет «съехавшие» фото исходного импорта.

Результат (рядом со скриптом):
  images/<код>.<ext>      — фото, переименованные по коду товара
  dahua_products.json     — данные для fix-dahua.mjs (код, цена, категория, имя, фото, флаг)
  dahua_mapping.csv       — то же в читаемом виде

Запуск (на сервере):
  python3 tools/dahua-fix/prepare.py --xlsx "/root/incoming/DAHUA PRICE new.xlsx"

Затем применить к БД:
  UPLOADS_DIR=/path/to/uploads/images node tools/dahua-fix/fix-dahua.mjs --apply --categories --names

Зависимость: openpyxl  (pip install openpyxl)
"""
import argparse, csv, json, os, re, zipfile
from xml.etree import ElementTree as ET
import openpyxl

NS = {'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
      'a':   'http://schemas.openxmlformats.org/drawingml/2006/main'}
EMB = '{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed'
PT = 12700  # EMU per point

CATMAP = [
    ("металлоискат", "Металлоискатели", "Металлоискатель"),
    ("видеодомофон", "Домофоны", "Видеодомофон"), ("домофон", "Домофоны", "Домофон"),
    ("вызывная wi", "Вызывные панели", "Вызывная панель"),
    ("вызывную панель", "Вызывные панели", "Вызывная панель"),
    ("вызывная панель", "Вызывные панели", "Вызывная панель"),
    ("монитор", "Мониторы", "Монитор"),
    ("видеорегистратор", "Видеорегистраторы", "Видеорегистратор"),
    ("видеомагнитофон", "Видеорегистраторы", "Видеорегистратор"),
    ("xvr", "Видеорегистраторы", "Видеорегистратор"), ("nvr", "Видеорегистраторы", "Видеорегистратор"),
    ("видеокамера", "Видеокамеры", "Видеокамера"), ("камера", "Видеокамеры", "Видеокамера"),
    ("объектив", "Объективы", "Объектив"),
    ("коммутатор", "Сетевое оборудование", "Коммутатор"), ("порт", "Сетевое оборудование", "Коммутатор"),
    ("switch", "Сетевое оборудование", "Коммутатор"),
    ("датчик", "Детекторы и датчики", "Датчик"), ("извещател", "Детекторы и датчики", "Извещатель"),
    ("видеодисплей", "Дисплеи", "Видеодисплей"), ("видеостен", "Дисплеи", "Видеодисплей"),
    ("дисплей", "Дисплеи", "Дисплей"), ("проектор", "Дисплеи", "Проектор"), ("экран", "Дисплеи", "Дисплей"),
    ("микрофон", "Аудио", "Микрофон"),
    ("сервер", "Серверы и ПО", "Сервер"), ("платформа", "Серверы и ПО", "Платформа"), ("dss", "Серверы и ПО", "ПО"),
    ("блок питания", "Питание", "Блок питания"), ("источник питания", "Питание", "Источник питания"),
    ("шлагбаум", "СКУД", "Шлагбаум"), ("турникет", "СКУД", "Турникет"), ("контроллер", "СКУД", "Контроллер"),
    ("считыватель", "СКУД", "Считыватель"), ("замок", "СКУД", "Замок"), ("панель", "Панели", "Панель"),
]

def categorize(desc):
    d = (desc or "").lower()
    for k, cat, sg in CATMAP:
        if k in d:
            return cat, sg
    return "Не определена", "Устройство"

def codenorm(code):
    c = (code or "").upper()
    c = re.sub(r'^DHI-', '', c)
    c = re.sub(r'^DH-', '', c)
    return re.sub(r'[^A-Z0-9]', '', c)

def safefile(code):
    return re.sub(r'[^A-Za-z0-9._-]', '_', code or "")

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default="/root/incoming/DAHUA PRICE new.xlsx", help="путь к прайс-листу")
    ap.add_argument("--out", default=os.path.dirname(os.path.abspath(__file__)), help="папка вывода")
    args = ap.parse_args()
    if not os.path.exists(args.xlsx):
        raise SystemExit(f"Не найден файл: {args.xlsx}")
    img_dir = os.path.join(args.out, "images")
    os.makedirs(img_dir, exist_ok=True)

    wb = openpyxl.load_workbook(args.xlsx, data_only=True)
    ws = wb.active
    default_h = ws.sheet_format.defaultRowHeight or 15.0
    maxr = ws.max_row
    heights = [default_h] * (maxr + 2)
    for r in range(1, maxr + 1):
        h = ws.row_dimensions[r].height
        if h:
            heights[r] = h
    codes = {r: (str(ws.cell(r, 1).value).strip() if ws.cell(r, 1).value else "") for r in range(1, maxr + 1)}
    prices = {r: ws.cell(r, 3).value for r in range(1, maxr + 1)}
    descs = {r: (str(ws.cell(r, 4).value).strip() if ws.cell(r, 4).value else "") for r in range(1, maxr + 1)}
    wb.close()

    cum = [0.0] * (maxr + 3)
    for r in range(2, maxr + 2):
        cum[r] = cum[r - 1] + heights[r - 1] * PT
    def row_of(y):
        for r in range(1, maxr + 1):
            if cum[r] <= y < cum[r + 1]:
                return r
        return maxr

    z = zipfile.ZipFile(args.xlsx)
    rels = z.read("xl/drawings/_rels/drawing1.xml.rels").decode()
    rid2media = {m.group(1): m.group(2).replace("../", "xl/")
                 for m in re.finditer(r'Id="([^"]+)"[^>]*Target="([^"]+)"', rels)}
    root = ET.fromstring(z.read("xl/drawings/drawing1.xml"))
    def g(e, t):
        x = e.find('xdr:' + t, NS)
        return int(x.text) if x is not None and x.text else 0
    imgs = []  # (center, top, bot, media)
    for an in root:
        frm = an.find('xdr:from', NS); to = an.find('xdr:to', NS); blip = an.find('.//a:blip', NS)
        if frm is None or blip is None:
            continue
        if g(frm, 'col') != 1:  # только колонка B
            continue
        media = rid2media.get(blip.get(EMB))
        if not media:
            continue
        top = cum[g(frm, 'row') + 1] + g(frm, 'rowOff')
        bot = (cum[g(to, 'row') + 1] + g(to, 'rowOff')) if to is not None else top + heights[g(frm, 'row') + 1] * PT
        imgs.append(((top + bot) / 2, top, bot, media))

    row2imgs = {}
    for c, t, b, m in imgs:
        row2imgs.setdefault(row_of(c), []).append((c, m))
    prod = [r for r in range(2, maxr + 1) if codes.get(r)]
    def rowcenter(r):
        return (cum[r] + cum[r + 1]) / 2
    def choose(r):
        lst = row2imgs.get(r, [])
        if len(lst) == 1:
            return lst[0][1], "ok"
        if len(lst) > 1:
            return sorted(lst)[0][1], "multi"
        rc = rowcenter(r)
        cover = sorted((abs((t + b) / 2 - rc), m) for c, t, b, m in imgs if t <= rc <= b)
        if cover:
            return cover[0][1], "covered"
        near = sorted((abs(c - rc), m) for c, t, b, m in imgs)
        return (near[0][1] if near else None), "nearest"

    records = []; saved = 0
    for r in prod:
        code = codes[r]
        media, flag = choose(r)
        ext = os.path.splitext(media)[1] if media else ".png"
        fname = safefile(code) + ext
        if media:
            open(os.path.join(img_dir, fname), "wb").write(z.read(media)); saved += 1
        cat, sg = categorize(descs[r])
        records.append({
            "row": r, "code": code, "code_norm": codenorm(code),
            "price_usd": round(prices[r], 4) if isinstance(prices[r], (int, float)) else prices[r],
            "category_guess": cat, "suggested_name": f"{sg} Dahua {code}",
            "image_file": fname, "image_flag": flag,
            "short_description": (descs[r].splitlines()[0][:200] if descs[r] else ""),
            "description": descs[r],
        })

    json.dump(records, open(os.path.join(args.out, "dahua_products.json"), "w"), ensure_ascii=False, indent=1)
    with open(os.path.join(args.out, "dahua_mapping.csv"), "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["row", "code", "code_norm", "price_usd", "category_guess", "suggested_name", "image_file", "flag", "short_description"])
        for x in records:
            w.writerow([x["row"], x["code"], x["code_norm"], x["price_usd"], x["category_guess"],
                        x["suggested_name"], x["image_file"], x["image_flag"], x["short_description"]])

    from collections import Counter
    print(f"Товаров: {len(records)} | фото сохранено: {saved}")
    print("Флаги фото:", dict(Counter(x["image_flag"] for x in records)))
    print("Категории:", dict(Counter(x["category_guess"] for x in records)))
    print(f"Готово -> {args.out}/dahua_products.json, images/, dahua_mapping.csv")

if __name__ == "__main__":
    main()
